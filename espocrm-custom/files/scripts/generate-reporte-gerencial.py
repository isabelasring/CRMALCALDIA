#!/usr/bin/env python3
"""Genera reporte gerencial en Excel o HTML (para PDF)."""

from __future__ import annotations

import argparse
import html
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None


def esc(value) -> str:
    return html.escape(str(value if value is not None else ""))


def logo_html(data: dict) -> str:
    b64 = data.get("logoBase64")
    mime = data.get("logoMime") or "image/png"

    if not b64:
        return ""

    return (
        f'<img src="data:{mime};base64,{b64}" '
        'alt="Alcaldía de Envigado" width="64" height="64" '
        'style="width:64px;height:64px;" />'
    )


def build_html(data: dict) -> str:
    kpis = data.get("kpis", {})
    filtro = data.get("filtroAsignado")

    th_style = (
        'bgcolor="#e2e8f0" style="border:1px solid #94a3b8;padding:5pt;'
        'font-family:Arial,sans-serif;font-size:9pt;color:#334155;"'
    )
    td_style = 'style="border:1px solid #94a3b8;padding:5pt;font-family:Arial,sans-serif;font-size:9pt;"'
    td_right = (
        'style="border:1px solid #94a3b8;padding:5pt;font-family:Arial,sans-serif;'
        'font-size:9pt;text-align:right;"'
    )
    td_num = td_right
    section_style = (
        'style="border:none;padding:10pt 0 4pt 0;font-family:Arial,sans-serif;'
        'font-size:10pt;font-weight:bold;color:#334155;"'
    )

    def semaforo_corto(label: str) -> str:
        cortos = {
            "Próximo a vencer": "Próx. a vencer",
            "Al día": "Al día",
            "Vencido": "Vencido",
        }
        return cortos.get(label, label)

    filtro_row = ""
    if filtro:
        filtro_row = (
            f"<tr><td colspan='2' {td_style}>"
            f"<b>Filtro:</b> Casos asignados a {esc(filtro)}</td></tr>"
        )

    meta_html = f"""<table width="100%" border="1" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
      <tr>
        <td width="50%" {td_style}><b>Generado:</b> {esc(data.get('generadoEn'))}</td>
        <td width="50%" {td_style}><b>Elaborado por:</b> {esc(data.get('generadoPor'))}</td>
      </tr>
      {filtro_row}
    </table>"""

    def summary_rows(rows: list[dict]) -> str:
        if not rows:
            return f"<tr><td colspan='2' {td_style} align='center'><i>Sin datos</i></td></tr>"
        out = ""
        for r in rows:
            out += (
                f"<tr>"
                f"<td width='80%' {td_style}>{esc(r.get('label', ''))}</td>"
                f"<td width='20%' {td_num}><b>{esc(r.get('total', 0))}</b></td>"
                f"</tr>"
            )
        return out

    def summary_block(title: str, label_col: str, rows: list[dict]) -> str:
        return f"""
    <table width="100%" border="0" cellspacing="0" cellpadding="0">
      <tr><td {section_style}>{esc(title)}</td></tr>
      <tr><td>
        <table width="100%" border="1" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
          <tr>
            <td width="80%" {th_style}><b>{esc(label_col)}</b></td>
            <td width="20%" {th_style} align="right"><b>Total</b></td>
          </tr>
          {summary_rows(rows)}
        </table>
      </td></tr>
    </table>"""

    detalle_rows = ""
    for row in data.get("detalle", []):
        semaforo_raw = row.get("semaforo") or ""
        semaforo = esc(semaforo_corto(semaforo_raw))
        sem_bg = "#ffffff"
        if semaforo_raw == "Vencido":
            sem_bg = "#fee2e2"
        elif semaforo_raw == "Próximo a vencer":
            sem_bg = "#fef3c7"
        elif semaforo_raw == "Al día":
            sem_bg = "#dcfce7"

        detalle_rows += (
            "<tr>"
            f'<td width="12%" {td_style}><font face="Courier New" size="1">{esc(row.get("radicado"))}</font></td>'
            f'<td width="16%" {td_style}>{esc(row.get("peticionario"))}</td>'
            f'<td width="10%" {td_style}>{esc(row.get("estado"))}</td>'
            f'<td width="9%" {td_style}>{esc(row.get("recurso"))}</td>'
            f'<td width="15%" {td_style}>{esc(row.get("asignado"))}</td>'
            f'<td width="10%" {td_style}>{esc(row.get("fechaVencimiento"))}</td>'
            f'<td width="18%" nowrap align="center" bgcolor="{sem_bg}" '
            f'style="border:1px solid #94a3b8;padding:5pt;font-family:Arial,sans-serif;'
            f'font-size:9pt;white-space:nowrap;">{semaforo}</td>'
            "</tr>"
        )

    if not detalle_rows:
        detalle_rows = f"<tr><td colspan='7' {td_style} align='center'><i>Sin casos</i></td></tr>"

    def kpi_cell(label: str, value, value_color: str = "#1e293b") -> str:
        return (
            f'<td width="33%" valign="top" style="border:1px solid #94a3b8;padding:8pt;'
            f'font-family:Arial,sans-serif;">'
            f'<font size="1" color="#64748b">{esc(label)}</font><br/>'
            f'<font size="4" color="{value_color}"><b>{esc(value)}</b></font>'
            f"</td>"
        )

    kpi_html = f"""<table width="100%" border="1" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
      <tr>
        {kpi_cell("Total casos", kpis.get("total", 0))}
        {kpi_cell("Pend. radicación", kpis.get("pendienteRadicacion", 0))}
        {kpi_cell("En gestión", kpis.get("enGestion", 0))}
      </tr>
      <tr>
        {kpi_cell("Finalizados", kpis.get("finalizados", 0))}
        {kpi_cell("Vencidos", kpis.get("vencidos", 0), "#b91c1c")}
        {kpi_cell("Próx. a vencer", kpis.get("proximosVencer", 0), "#a16207")}
      </tr>
    </table>"""

    summaries = (
        summary_block("Resumen por estado", "Estado", data.get("porEstado", []))
        + summary_block("Resumen por recurso / tema", "Recurso", data.get("porRecurso", []))
        + summary_block("Resumen por canal de reporte", "Canal", data.get("porCanal", []))
        + summary_block("Semáforo de vencimiento", "Semáforo", data.get("porSemaforo", []))
    )

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8" />
  <meta http-equiv="Content-Type" content="text/html; charset=utf-8" />
  <title>{esc(data.get('titulo', 'Reporte gerencial'))}</title>
</head>
<body style="margin:1.5cm 1.5cm 1.5cm 1.5cm;font-family:Arial,sans-serif;color:#1e293b;font-size:10pt;">

  <table width="100%" border="0" cellspacing="0" cellpadding="0">
    <tr>
      <td width="70" valign="middle">{logo_html(data)}</td>
      <td valign="middle" style="padding-left:10pt;">
        <font face="Arial" size="4" color="#1e293b"><b>{esc(data.get('titulo'))}</b></font><br/>
        <font face="Arial" size="2" color="#64748b">{esc(data.get('subtitulo'))}</font>
      </td>
    </tr>
    <tr><td colspan="2" style="border-bottom:2px solid #cbd5e1;height:8pt;">&nbsp;</td></tr>
  </table>

  <br/>

  {meta_html}

  <br/>

  <table width="100%" border="0" cellspacing="0" cellpadding="0">
    <tr><td {section_style}>Indicadores generales</td></tr>
    <tr><td>{kpi_html}</td></tr>
  </table>

  <br/>

  {summaries}

  <br/>

  <table width="100%" border="0" cellspacing="0" cellpadding="0">
    <tr><td {section_style}>Detalle de casos</td></tr>
    <tr><td>
      <table width="100%" border="1" cellspacing="0" cellpadding="0" style="border-collapse:collapse;table-layout:fixed;">
        <colgroup>
          <col width="12%"/><col width="16%"/><col width="10%"/><col width="9%"/>
          <col width="15%"/><col width="10%"/><col width="18%"/>
        </colgroup>
        <tr>
          <td {th_style}><b>Radicado</b></td>
          <td {th_style}><b>Peticionario</b></td>
          <td {th_style}><b>Estado</b></td>
          <td {th_style}><b>Recurso</b></td>
          <td {th_style}><b>Asignado</b></td>
          <td {th_style}><b>Vencimiento</b></td>
          <td {th_style} align="center"><b>Semáforo</b></td>
        </tr>
        {detalle_rows}
      </table>
    </td></tr>
  </table>

  <br/>

  <table width="100%" border="0" cellspacing="0" cellpadding="0">
    <tr>
      <td align="center" style="border-top:1px solid #cbd5e1;padding-top:6pt;">
        <font face="Arial" size="1" color="#94a3b8">
          Alcaldía de Envigado — Secretaría de Medio Ambiente · CRM ambiental
        </font>
      </td>
    </tr>
  </table>

</body>
</html>"""


def write_xlsx(data: dict, output_path: Path) -> None:
    if Workbook is None:
        print("Falta openpyxl.", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    header_fill = PatternFill("solid", fgColor="E2E8F0")
    header_font = Font(color="334155", bold=True)

    ws["A1"] = data.get("titulo", "Reporte gerencial")
    ws["A1"].font = Font(size=14, bold=True, color="1E293B")
    ws["A2"] = data.get("subtitulo", "")
    ws["A3"] = f"Generado: {data.get('generadoEn', '')}"
    ws["A4"] = f"Elaborado por: {data.get('generadoPor', '')}"

    row = 6
    ws.cell(row=row, column=1, value="Indicador").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2, value="Total").fill = header_fill
    ws.cell(row=row, column=2).font = header_font

    kpis = data.get("kpis", {})
    indicadores = [
        ("Total casos", kpis.get("total", 0)),
        ("Pendientes de radicación", kpis.get("pendienteRadicacion", 0)),
        ("En gestión", kpis.get("enGestion", 0)),
        ("Finalizados", kpis.get("finalizados", 0)),
        ("Vencidos (activos)", kpis.get("vencidos", 0)),
        ("Próximos a vencer", kpis.get("proximosVencer", 0)),
    ]

    for label, total in indicadores:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=total)

    def write_sheet(name: str, rows: list[dict]) -> None:
        sheet = wb.create_sheet(name[:31])
        sheet.append([name.split(" ", 1)[0], "Total"])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for item in rows:
            sheet.append([item.get("label", ""), item.get("total", 0)])

    write_sheet("Por estado", data.get("porEstado", []))
    write_sheet("Por recurso", data.get("porRecurso", []))
    write_sheet("Por canal", data.get("porCanal", []))
    write_sheet("Por semaforo", data.get("porSemaforo", []))

    detalle = wb.create_sheet("Detalle casos")
    detalle.append([
        "Radicado", "Expediente", "Peticionario", "Estado", "Recurso", "Canal",
        "Asignado", "Fecha caso", "Vencimiento", "Semáforo", "Barrio",
    ])
    for cell in detalle[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in data.get("detalle", []):
        detalle.append([
            item.get("radicado"),
            item.get("expediente"),
            item.get("peticionario"),
            item.get("estado"),
            item.get("recurso"),
            item.get("canal"),
            item.get("asignado"),
            item.get("fechaCaso"),
            item.get("fechaVencimiento"),
            item.get("semaforo"),
            item.get("barrio"),
        ])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(max_len + 2, 42)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload", required=True)
    parser.add_argument("--format", required=True, choices=["pdf", "xlsx"])
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    payload_path = Path(args.payload)
    output_base = Path(args.output)

    with payload_path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)

    if args.format == "xlsx":
        write_xlsx(data, output_base.with_suffix(".xlsx"))
        return

    html_content = build_html(data)
    output_base.with_suffix(".html").write_text(html_content, encoding="utf-8")


if __name__ == "__main__":
    main()
